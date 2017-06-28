import copy
import os
from datetime import datetime

import re
from flask import session, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import Worksheet


class Excel(object):

    def __init__(self, template_file=None):
        if template_file:
            # テンプレートファイルの読み込み
            self.workbook = load_workbook("excel/templates/" + template_file)
        else:
            # 新規ファイルの作成
            self.workbook = Workbook()
        directory = datetime.today().strftime("%Y%m%d")
        path = "excel/temporary/" + directory
        hms = datetime.today().strftime("%H%M%S") + "%03d" % (datetime.now().microsecond // 1000)
        temp_file_name = hms + "temp" + "_" + str(session['user']['id'])
        if not os.path.isdir(path):
            os.makedirs(path)
        self.file_name = template_file
        self.file_path = path + "/" + temp_file_name

    def save(self, file_name=None):
        if file_name:
            self.file_name = file_name
        self.workbook.save(self.file_path)

    def download(self):
        return send_file("../" + self.file_path,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         attachment_filename=self.file_name, as_attachment=True)

    @property
    def active(self):
        return self.workbook.active

    @active.setter
    def active(self, value):
        self.workbook.active = value


def insert_rows(self, row_idx, cnt, above=False, copy_style=True, fill_formulae=True):
    """Inserts new (empty) rows into worksheet at specified row index.

    :param self: 
    :param row_idx: Row index specifying where to insert new rows.
    :param cnt: Number of rows to insert.
    :param above: Set True to insert rows above specified row index.
    :param copy_style: Set True if new rows should copy style of immediately above row.
    :param fill_formulae: Set True if new rows should take on formula from immediately above row, 
                          filled with references new to rows.

    Usage:

    * insert_rows(2, 10, above=True, copy_style=False)

    """
    cell_re = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    row_idx = row_idx - 1 if above else row_idx

    def replace(m):
        row_ = m.group('row')
        prefix = "$" if row_.find("$") != -1 else ""
        row_ = int(row_.replace("$", ""))
        row_ += cnt if row_ > row_idx else 0
        return m.group('col') + prefix + str(row_)

    # First, we shift all cells down cnt rows...
    old_cells = set()
    old_fas = set()
    new_cells = dict()
    new_fas = dict()
    for c in self._cells.values():

        old_coor = c.coordinate

        # Shift all references to anything below row_idx
        if c.data_type == Cell.TYPE_FORMULA:
            c.value = cell_re.sub(
                replace,
                c.value
            )
            # Here, we need to properly update the formula references to reflect new row indices
            if old_coor in self.formula_attributes and 'ref' in self.formula_attributes[old_coor]:
                self.formula_attributes[old_coor]['ref'] = cell_re.sub(
                    replace,
                    self.formula_attributes[old_coor]['ref']
                )

        # Do the magic to set up our actual shift
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row, c.col_idx))
            c.row += cnt
            new_cells[(c.row, c.col_idx)] = c
            if old_coor in self.formula_attributes:
                old_fas.add(old_coor)
                fa = self.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa

    for coor in old_cells:
        del self._cells[coor]
    self._cells.update(new_cells)

    for fa in old_fas:
        del self.formula_attributes[fa]
    self.formula_attributes.update(new_fas)

    # Next, we need to shift all the Row Dimensions below our new rows down by cnt...
    for row in range(len(self.row_dimensions)-1+cnt, row_idx+cnt, -1):
        new_rd = copy.copy(self.row_dimensions[row-cnt])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        del self.row_dimensions[row-cnt]

    # Now, create our new rows, with all the pretty cells
    row_idx += 1
    for row in range(row_idx, row_idx+cnt):
        # Create a Row Dimension for our new row
        new_rd = copy.copy(self.row_dimensions[row-1])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        for col in range(1, self.max_column):
            col = get_column_letter(col)
            cell = self.cell('%s%d' % (col, row))
            cell.value = None
            source = self.cell('%s%d' % (col, row-1))
            if copy_style:
                cell.number_format = source.number_format
                cell.font = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border = source.border.copy()
                cell.fill = source.fill.copy()
            if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
                s_coor = source.coordinate
                if s_coor in self.formula_attributes and 'ref' not in self.formula_attributes[s_coor]:
                    fa = self.formula_attributes[s_coor].copy()
                    self.formula_attributes[cell.coordinate] = fa
                # print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
                cell.value = re.sub(
                    "(\$?[A-Z]{1,3}\$?)%d" % (row - 1),
                    lambda m: m.group(1) + str(row),
                    source.value
                )
                cell.data_type = Cell.TYPE_FORMULA

    # Check for Merged Cell Ranges that need to be expanded to contain new cells
    for cr_idx, cr in enumerate(self.merged_cell_ranges):
        self.merged_cell_ranges[cr_idx] = cell_re.sub(
            replace,
            cr
        )

Worksheet.insert_rows = insert_rows
