from datetime import datetime

from openpyxl.styles import Border, Side, Font

from application.domain.model.immutables.billing_timing import BillingTiming
from application.domain.repository.excel import Excel


class ClientOrderReport(object):

    def __init__(self, project):
        self.project = project
        self.excel = Excel("client_order.xlsx")
        self.ws = self.excel.active

    def download(self):
        self._create_excel()

        return self.excel.download()

    def _create_excel(self):
        self.write_client_order().write_style()

        # エクセルを一時フォルダに保存
        self.excel.save('04_注文請書（{}）_{}.xlsx'.format(
                                                       self.project.client_order_no or "",
                                                       datetime.today().strftime("%Y%m%d")))

    def write_client_order(self):
        # 値を代入
        self.ws[self.excel.get_defined_name_range("printed_date")].value = datetime.today().date()
        if self.project.client_company:
            self.ws[self.excel.get_defined_name_range("client_company_name")].value = \
                self.project.client_company.company_name
        self.ws[self.excel.get_defined_name_range("client_order_no")].value = self.project.client_order_no
        self.ws[self.excel.get_defined_name_range("estimation_no")].value = self.project.estimation_no
        self.ws[self.excel.get_defined_name_range("project_name")].value = self.project.project_name
        if self.project.contract_form:
            self.ws[self.excel.get_defined_name_range("contract_form")].value = self.project.contract_form.name
        self.ws[self.excel.get_defined_name_range("estimated_total_amount")].value = \
            self.project.estimated_total_amount or 0
        self.ws[self.excel.get_defined_name_range("contents")].value = self.project.contents
        self.ws[self.excel.get_defined_name_range("start_date")].value = self.project.start_date
        self.ws[self.excel.get_defined_name_range("end_date")].value = self.project.end_date
        self.ws[self.excel.get_defined_name_range("responsible_person")].value = self.project.responsible_person
        self.ws[self.excel.get_defined_name_range("quality_control")].value = self.project.quality_control
        self.ws[self.excel.get_defined_name_range("subcontractor")].value = self.project.subcontractor
        self.ws[self.excel.get_defined_name_range("working_place")].value = self.project.working_place
        self.ws[self.excel.get_defined_name_range("delivery_place")].value = self.project.delivery_place
        self.ws[self.excel.get_defined_name_range("deliverables")].value = self.project.deliverables
        if self.project.billing_timing == BillingTiming.billing_by_month:
            self.ws[self.excel.get_defined_name_range("inspection_date")].value = '毎月月末'
        else:
            self.ws[self.excel.get_defined_name_range("inspection_date")].value = self.project.inspection_date
        if self.project.billing_timing:
            self.ws[self.excel.get_defined_name_range("billing_timing")].value = \
                self.project.billing_timing.name_for_report
        self.ws[self.excel.get_defined_name_range("remarks")].value = self.remarks()
        return self

    def write_style(self):
        # フォント設定
        self.ws[self.excel.get_defined_name_range("client_company_name")].font = \
            Font(name="ＭＳ ゴシック", size=14, underline="single")
        # 表示形式
        self.ws[self.excel.get_defined_name_range("printed_date")].number_format = 'yyyy年m月d日'
        self.ws[self.excel.get_defined_name_range("start_date")].number_format = 'yyyy年m月d日'
        self.ws[self.excel.get_defined_name_range("end_date")].number_format = 'yyyy年m月d日'
        if self.project.billing_timing == BillingTiming.billing_at_last:
            self.ws[self.excel.get_defined_name_range("inspection_date")].number_format = 'yyyy"年"m"月"d"日"'
        # 罫線
        self.write_border_to_merged_cell(row=14)
        self.write_border_to_merged_cell(row=16)
        self.write_border_to_cell(row=18)
        self.write_border_to_cell(row=19)
        self.write_border_to_cell(row=20)
        self.write_border_to_merged_cell(row=21)
        self.write_border_to_cell(row=23)
        self.write_border_to_cell(row=24)
        self.write_border_to_cell(row=25)
        self.write_border_to_cell(row=26)
        self.write_border_to_cell(row=27)
        self.write_border_to_cell(row=28)
        self.write_border_to_cell(row=29)
        self.write_border_to_cell(row=30)
        self.write_border_to_merged_cell(row=31)
        self.ws['B32'].border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

    # セル結合されていない1行の項目に罫線を引く
    def write_border_to_cell(self, row):
        for column_num in ['D', 'E', 'F', 'G']:
            self.ws[column_num + str(row)].border = Border(bottom=Side(style='thin'))
        self.ws['H' + str(row)].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

    # セルが結合されて2行になっている項目に罫線を引く
    def write_border_to_merged_cell(self, row):
        self.ws['H' + str(row)].border = Border(top=Side(style='thin'), right=Side(style='thin'))
        for column_num in ['D', 'E', 'F', 'G']:
            self.ws[column_num + str(row + 1)].border = Border(bottom=Side(style='thin'))
        self.ws['H' + str(row + 1)].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

    def remarks(self):
        text = '上記以外の条件は当社見積書(見積書No.{})の通りとします。'\
            .format(self.project.estimation_no or "               ")
        return text
