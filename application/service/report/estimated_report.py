import locale
from datetime import datetime

from erajp.converter import strjpftime
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font, Alignment

from application.domain.model.immutables.billing_timing import BillingTiming
from application.domain.model.immutables.contract import Contract
from application.domain.repository.excel import Excel


class EstimatedReport(object):

    def __init__(self, project):
        self.project = project
        self.excel = Excel("estimation.xlsx")
        self.current_row = 22
        self.ws = self.excel.active

    def download(self):
        self._create_excel()

        return self.excel.download()

    def _create_excel(self):
        # 上部部分記載
        self.write_estimated_content_rows()
        # 作業詳細記載
        self.create_project_detail_rows()
        # 作業内容～備考
        self.create_contract_content_rows()

        # ロゴ画像貼り付け
        img = Image('excel/templates/tp_logo.png')
        img.width = img.width * 0.63
        img.height = img.height * 0.63
        self.ws.add_image(img, 'I10')

        # 余白調整
        self.ws.page_margins.top = 0.7
        self.ws.page_margins.left = 0.3
        self.ws.page_margins.right = 0.26
        self.ws.page_margins.bottom = 0.2
        self.ws.page_margins.header = 0.7
        self.ws.page_margins.footer = 0.22
        # 倍率の調整
        self.ws.page_setup.scale = 86

        # エクセルを一時フォルダに保存
        self.excel.save('01_見積書（{}）_{}.xlsx'.format(self.project.estimation_no,
                                                    datetime.today().strftime("%Y%m%d")))

    def write_estimated_content_rows(self):
        # 結合しているセルの罫線が消えるため、罫線を引き直す。
        for rows in self.ws[self.excel.get_defined_name_range("estimated_content")]:
            for cell in rows:
                cell.border = Border(outline=True, bottom=Side(style='dashed'))
        # 値を代入
        self.ws[self.excel.get_defined_name_range("estimation_no")].value = self.project.estimation_no
        self.ws[self.excel.get_defined_name_range("printed_date")].value = datetime.today().date()
        if self.project.client_company:
            self.ws[self.excel.get_defined_name_range("client_company_name")].value = \
                self.project.client_company.company_name
        self.ws[self.excel.get_defined_name_range("project_name")].value = self.project.project_name
        # start_date、end_dateをdatetime型に変更
        start_date = datetime(self.project.start_date.year, self.project.start_date.month, self.project.start_date.day)
        end_date = datetime(self.project.end_date.year, self.project.end_date.month, self.project.end_date.day)
        locale.setlocale(locale.LC_ALL, '')
        self.ws[self.excel.get_defined_name_range("start_date")].value = strjpftime(start_date, '  %O%E年%m月%d日')
        self.ws[self.excel.get_defined_name_range("end_date")].value = strjpftime(end_date, '  %O%E年%m月%d日')
        if self.project.billing_timing:
            self.ws[self.excel.get_defined_name_range("billing_timing")].value = \
                self.project.billing_timing.name_for_report
        if self.project.contract_form:
            self.ws[self.excel.get_defined_name_range("contract_form")].value = self.project.contract_form.name
        # 表示形式
        self.ws[self.excel.get_defined_name_range("printed_date")].number_format = 'yyyy年m月d日'
        # 請負ではない場合、「瑕疵担保期間」を非表示にする。
        if self.project.contract_form != Contract.blanket:
            self.ws.row_dimensions[19].hidden = True
            # 書式が崩れるので修正
            self.ws['J18'].border = Border(bottom=Side(style='medium'))
            self.ws['L18'].border = Border(bottom=Side(style='medium'))

    def create_project_detail_rows(self):
        # 作業内容タイトル-------------------------------------------------------------------------
        for column_num in range(2, 14):
            self.ws.cell(row=self.current_row, column=column_num).border = Border(bottom=Side(style='medium'))
        self.current_row += 1
        self.create_project_detail_line(bottom_line_style='double')
        self.current_row += 1

        # 作業内容--------------------------------------------------------------------------------
        for n, project_detail in enumerate(self.project.project_details):
            # 書式作成
            self.create_project_detail_style()
            # 値代入
            self.ws['B' + str(self.current_row)].value = (n + 1)
            if project_detail.engineer_id:
                self.ws['C' + str(self.current_row)].value = project_detail.engineer.engineer_name
            else:
                self.ws['C' + str(self.current_row)].value = project_detail.work_name
            self.ws['C' + str(self.current_row)].font = Font(name="ＭＳ ゴシック", size=9, bold=False)
            self.ws['G' + str(self.current_row)].value = project_detail.billing_money
            self.ws['H' + str(self.current_row)].value = project_detail.remarks
            # 表示形式
            self.ws['G' + str(self.current_row)].number_format = '¥#,###.-'
            self.current_row += 1

        # 消費税----------------------------------------------------------------------------------
        # 書式作成
        self.create_project_detail_style()
        # 値代入
        if self.project.billing_tax:
            tax = str(self.project.client_company.billing_tax.name)
        else:
            tax = ''
        self.ws['C' + str(self.current_row)].value = '消費税（' + tax + '）'
        self.ws['G' + str(self.current_row)].value = self.project.tax_of_estimated_total_amount()
        # 表示形式
        self.ws['C' + str(self.current_row)].number_format = '"  "@'
        self.ws['G' + str(self.current_row)].number_format = '¥#,###.-'
        self.current_row += 1

        # 合計金額--------------------------------------------------------------------------------
        # 書式設定
        self.create_project_detail_style()
        # 値代入
        self.ws['C' + str(self.current_row)].value = '合計'
        self.ws['G' + str(self.current_row)].value = (self.project.estimated_total_amount or 0)\
            + self.project.tax_of_estimated_total_amount()
        # 上部に存在する「金額」は、合計金額を参照するようにする。
        self.ws['E14'].value = '=G' + str(self.current_row)
        # 表示形式
        self.ws['C' + str(self.current_row)].number_format = '"  "@'
        self.ws['G' + str(self.current_row)].number_format = '¥#,###.-'
        self.current_row += 1

        # その他調整------------------------------------------------------------------------------
        # 空白行を作成
        if len(self.project.project_details) < 3:
            for n in range(3-len(self.project.project_details)):
                # 書式作成
                self.create_project_detail_style()
                self.current_row += 1

    def create_project_detail_style(self):
        # セルの結合
        self.ws.merge_cells('C' + str(self.current_row) + ':F' + str(self.current_row))
        self.ws.merge_cells('H' + str(self.current_row) + ':M' + str(self.current_row))
        # フォント設定
        self.ws['B' + str(self.current_row)].font = Font(name="ＭＳ ゴシック", bold=True)
        self.ws['C' + str(self.current_row)].font = Font(name="ＭＳ ゴシック", bold=False)
        self.ws['G' + str(self.current_row)].font = Font(name="ＭＳ ゴシック", bold=False)
        self.ws['H' + str(self.current_row)].font = Font(name="ＭＳ ゴシック", bold=False)
        # 書式設定
        self.ws['B' + str(self.current_row)].alignment = Alignment(horizontal='center')
        self.ws['C' + str(self.current_row)].alignment = Alignment(wrap_text=True)
        self.ws['G' + str(self.current_row)].alignment = Alignment(horizontal='right')
        self.ws['H' + str(self.current_row)].alignment = Alignment(wrap_text=True)
        # 罫線
        self.create_project_detail_line(bottom_line_style='thin')

    def create_project_detail_line(self, bottom_line_style):
        self.ws['B' + str(self.current_row)].border = Border(
                                                        left=Side(style='medium'),
                                                        right=Side(style='thin'),
                                                        bottom=Side(style=bottom_line_style))
        for column_num in ['C', 'D', 'E', 'F']:
            self.ws[column_num + str(self.current_row)].border = Border(bottom=Side(style=bottom_line_style))
        self.ws['G' + str(self.current_row)].border = Border(
                                                        left=Side(style='thin'),
                                                        right=Side(style='double'),
                                                        bottom=Side(style=bottom_line_style))
        for column_num in ['H', 'I', 'J', 'K', 'L']:
            self.ws[column_num + str(self.current_row)].border = Border(bottom=Side(style=bottom_line_style))
        self.ws['M' + str(self.current_row)].border = Border(
                                                        right=Side(style='medium'),
                                                        bottom=Side(style=bottom_line_style))

    def create_contract_content_rows(self):
        self.create_contract_contents_row('作業内容', self.project.contents, top_line_style='double')
        self.create_contract_contents_row('納品物', self.project.deliverables)
        self.create_contract_contents_row('作業場所', self.project.working_place)
        if self.project.billing_timing == BillingTiming.billing_by_month:
            self.create_contract_contents_row('検査完了日', '毎月月末')
        else:
            self.create_contract_contents_row('検査完了日', self.project.inspection_date)
        self.create_contract_contents_row('作業責任者', self.project.responsible_person)
        self.create_contract_contents_row('品質管理担当者', self.project.quality_control)
        self.create_contract_contents_row('再委託先', self.project.subcontractor)
        self.create_contract_contents_row('備考', self.project.remarks)

    def create_contract_contents_row(self, title, value, top_line_style=None):
        # セルの結合
        self.ws.merge_cells('B' + str(self.current_row) + ':C' + str(self.current_row))
        self.ws.merge_cells('E' + str(self.current_row) + ':M' + str(self.current_row))
        # フォント設定
        self.ws['B' + str(self.current_row)].font = Font(name="ＭＳ ゴシック", bold=True)
        self.ws['E' + str(self.current_row)].font = Font(name="ＭＳ ゴシック", bold=False)
        # 書式設定
        self.ws['B' + str(self.current_row)].alignment = Alignment(
                                                            horizontal='distributed',
                                                            vertical='center',
                                                            wrap_text=True,
                                                            justifyLastLine=1)
        self.ws['E' + str(self.current_row)].alignment = Alignment(
                                                            horizontal='left',
                                                            vertical='center',
                                                            wrap_text=True)
        # 罫線
        self.ws['B' + str(self.current_row)].border = Border(
                                                        top=Side(style=top_line_style),
                                                        left=Side(style='medium'),
                                                        bottom=Side(style='thin'))
        self.ws['C' + str(self.current_row)].border = Border(
                                                        top=Side(style=top_line_style),
                                                        bottom=Side(style='thin'))
        self.ws['D' + str(self.current_row)].border = Border(
                                                        top=Side(style=top_line_style),
                                                        left=Side(style='thin'),
                                                        bottom=Side(style='thin'))
        for column_num in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            self.ws[column_num + str(self.current_row)].border = Border(
                                                                    top=Side(style=top_line_style),
                                                                    bottom=Side(style='thin'))
        self.ws['M' + str(self.current_row)].border = Border(
                                                        top=Side(style=top_line_style),
                                                        right=Side(style='medium'),
                                                        bottom=Side(style='thin'))
        # 値を代入
        self.ws['B' + str(self.current_row)].value = title
        self.ws['E' + str(self.current_row)].value = value
        # 表示形式
        if title == '検査完了日':
            self.ws['E' + str(self.current_row)].number_format = 'yyyy年m月d日'
        self.current_row += 1
