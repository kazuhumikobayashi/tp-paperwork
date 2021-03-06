from datetime import datetime

from erajp.converter import strjpftime
from openpyxl.styles import Border, Side, Font, Alignment

from application.domain.model.immutables.rule import Rule
from application.domain.repository.engineer_history_repository import EngineerHistoryRepository
from application.domain.repository.excel import Excel
from application.service.report.sheet.address_sheet import AddressSheet


class BpOrderReport(object):
    engineer_history_repository = EngineerHistoryRepository()

    def __init__(self, project_detail):
        self.project_detail = project_detail
        self.excel = Excel("bp_order_report.xlsx")
        self.address_sheet = AddressSheet(self)

    def download(self):
        self._create_excel()

        return self.excel.download()

    def _create_excel(self):
        self.ws = self.excel.workbook['注文書']
        self.write_bp_order_original().create_bp_order_style()

        self.ws = self.excel.workbook['注文書_副']
        self.create_bp_order_style()

        self.ws = self.excel.workbook['注文請書']
        self.write_bp_order_confirmation().create_bp_order_style()
        # 宛名を作成
        self.address_sheet.create_address_sheet()

        # エクセルを一時フォルダに保存
        self.excel.save('02_注文書（{}_{}）_{}.xlsx'.format(
                                                       self.project_detail.engineer.engineer_name,
                                                       self.project_detail.bp_order_no,
                                                       datetime.today().strftime("%Y%m%d")))

    def write_bp_order_original(self):
        # 値を代入
        engineer_history = self.engineer_history_repository.get_history_by_date(self.project_detail.engineer.id,
                                                                                self.project_detail.billing_start_day)
        self.ws[self.excel.get_defined_name_range("bp_order_no")].value = self.project_detail.bp_order_no
        self.ws[self.excel.get_defined_name_range("printed_date")].value = self.project_detail.billing_start_day
        self.ws[self.excel.get_defined_name_range("bp_company_name")].value = \
            self.project_detail.engineer.company.company_name
        if self.project_detail.engineer.company.contract_date:
            self.ws[self.excel.get_defined_name_range("contract_date")].value = \
                strjpftime(datetime(
                    self.project_detail.engineer.company.contract_date.year,
                    self.project_detail.engineer.company.contract_date.month,
                    self.project_detail.engineer.company.contract_date.day
                ), '%O%E年%m月%d日')
        self.ws[self.excel.get_defined_name_range("project_name_for_bp")].value = \
            self.project_detail.project.project_name_for_bp
        self.ws[self.excel.get_defined_name_range("start_date")].value = self.project_detail.billing_start_day
        self.ws[self.excel.get_defined_name_range("end_date")].value = self.project_detail.billing_end_day
        if engineer_history:
            self.ws[self.excel.get_defined_name_range("payment_per_month")].value = engineer_history.payment_per_month
            self.ws[self.excel.get_defined_name_range("payment_detail")].value = \
                self.get_payment_detail_text(engineer_history)
            self.ws[self.excel.get_defined_name_range("payment_condition")].value = engineer_history.payment_condition
            self.ws[self.excel.get_defined_name_range("remarks")].value = engineer_history.remarks
        return self

    def write_bp_order_confirmation(self):
        # 印紙の罫線を修正
        self.ws['B3'].border = Border(left=Side(style='hair'), right=Side(style='hair'))
        self.ws['B4'].border = Border(left=Side(style='hair'), right=Side(style='hair'), bottom=Side(style='hair'))
        return self

    def create_bp_order_style(self):
        bp_company_name = self.ws['B9']
        contract_date = self.ws['B18']
        printed_date = self.ws['J4']
        start_date = self.ws['D26']
        end_date = self.ws['D27']
        payment_per_month = self.ws['D31']

        # フォント
        if self.ws.title == '注文書' or self.ws.title == '注文書_副':
            bp_company_name.font = Font(name='ＭＳ 明朝', size='10.5', underline='single')

        # 書式設定
        contract_date.alignment = Alignment(horizontal='right')
        payment_per_month.alignment = Alignment(horizontal='left')

        # 表示形式
        printed_date.number_format = 'yyyy"年"m"月"d"日"'
        contract_date.number_format = 'yyyy"年"m"月"d"日"'
        start_date.number_format = 'yyyy"年"m"月"d"日"'
        end_date.number_format = 'yyyy"年"m"月"d"日"'
        payment_per_month.number_format = '¥#,##0.-"/月額"'

        # 罫線
        for column_num in ['C', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            self.ws[column_num + str(37)].border = Border(bottom=Side(style='medium'))

        # 倍率の調整
        self.ws.page_setup.scale = 91
        return self

    def get_payment_detail_text(self, engineer_history):
        text = self.project_detail.engineer.engineer_name + " 氏  "\
                 + "¥{:,d}.-/月額".format(engineer_history.payment_per_month)
        if engineer_history.payment_rule == Rule.variable:
            text += "\n        "\
                 + "超過単価：" + "¥{:,d}.-/H".format(engineer_history.payment_per_top_hour) + "  "\
                 + "欠業単価：" + "¥{:,d}.-/H".format(engineer_history.payment_per_bottom_hour)
        else:
            text += "（固定）"
        return text

    def get_print_name(self):
        return self.project_detail.engineer.company.print_name
