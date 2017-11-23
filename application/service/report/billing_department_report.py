from datetime import datetime

from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

from application.domain.repository.excel import Excel


class BillingDepartmentReport(object):

    DEPARTMENT = 'department'
    DEPOSIT = 'deposit'
    
    def __init__(self, project_months, month):
        self.project_months = project_months
        self.month = month
        self.excel = Excel("billing_department.xlsx")
        self.current_row = 3

    def download(self):
        self._create_excel_department()
        self._create_excel_deposit()

        return self.excel.download()

    def _create_excel_department(self):
        self.ws = self.excel.workbook['Sheet1']

        group_name = ""
        department_name = ""
        company_name = ""
        # "顧客"単位の各合計金額の宣言
        transportation_total_company = 0
        tax_transportation_total_company = 0
        confirmation_total_company = 0
        tax_total_company = 0
        amount_total_company = 0
        # "部"単位の各合計金額の宣言
        transportation_total_department = 0
        tax_transportation_total_department = 0
        confirmation_total_department = 0
        tax_total_department = 0
        amount_total_department = 0
        # "本部"単位の各合計金額の宣言
        transportation_total_group = 0
        tax_transportation_total_group = 0
        confirmation_total_group = 0
        tax_total_group = 0
        amount_total_group = 0
        # 全合計金額の宣言
        transportation_total_all = 0
        tax_transportation_total_all = 0
        confirmation_total_all = 0
        tax_total_all = 0
        amount_total_all = 0
        i = 0
        for i, project_month in enumerate(self.project_months):
                
            if i != 0:
                # １つ上の"顧客"の値を比較
                if company_name != project_month.project.client_company.company_name:
                    self.amount_total(i, 'company', company_name, transportation_total_company,
                                      tax_transportation_total_company, confirmation_total_company,
                                      tax_total_company, amount_total_company)
                    transportation_total_company = 0
                    tax_transportation_total_company = 0
                    confirmation_total_company = 0
                    tax_total_company = 0
                    amount_total_company = 0
                # １つ上の"部"の値を比較
                if department_name != project_month.project.recorded_department.department_name:
                    self.amount_total(i, 'department', department_name, transportation_total_department,
                                      tax_transportation_total_department, confirmation_total_department,
                                      tax_total_department, amount_total_department)
                    transportation_total_department = 0
                    tax_transportation_total_department = 0
                    confirmation_total_department = 0
                    tax_total_department = 0
                    amount_total_department = 0
                # １つ上の"本部"の値を比較
                if group_name != project_month.project.recorded_department.group_name:
                    self.amount_total(i, 'group', group_name, transportation_total_group,
                                      tax_transportation_total_group, confirmation_total_group, tax_total_group,
                                      amount_total_group)
                    transportation_total_group = 0
                    tax_transportation_total_group = 0
                    confirmation_total_group = 0
                    tax_total_group = 0
                    amount_total_group = 0
            # 値を代入
            self.ws['B' + str(i + self.current_row)].value = project_month.project.recorded_department.group_name
            self.ws['C' + str(i + self.current_row)].value = project_month.project.recorded_department.department_name
            self.ws['D' + str(i + self.current_row)].value = project_month.project.client_company.company_name
            self.ws['E' + str(i + self.current_row)].value = project_month.client_billing_no
            self.ws['F' + str(i + self.current_row)].value = project_month.billing_printed_date
            self.ws['G' + str(i + self.current_row)].value = project_month.project.project_name
            self.ws['H' + str(i + self.current_row)].value = project_month.billing_transportation or 0
            self.ws['I' + str(i + self.current_row)].value = project_month.get_tax_of_billing_transportation()
            self.ws['J' + str(i + self.current_row)].value = project_month.billing_confirmation_money \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            self.ws['K' + str(i + self.current_row)].value = project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            self.ws['L' + str(i + self.current_row)].value = (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)
            self.ws['M' + str(i + self.current_row)].value = project_month.deposit_date
            self.ws['N' + str(i + self.current_row)].value = project_month.project.client_company.bank.bank_name
            # 金額セルの書式設定
            self.money_format(i, 'department')
            # 日付セルの書式設定
            self.day_format(i, 'department')
            # "顧客"単位の各合計金額
            transportation_total_company += project_month.billing_transportation or 0
            tax_transportation_total_company += project_month.get_tax_of_billing_transportation()
            confirmation_total_company += project_month.billing_confirmation_money \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            tax_total_company += project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            amount_total_company += (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)
            # "部"単位の各合計金額
            transportation_total_department += project_month.billing_transportation or 0
            tax_transportation_total_department += project_month.get_tax_of_billing_transportation()
            confirmation_total_department += project_month.billing_confirmation_money \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            tax_total_department += project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            amount_total_department += (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)
            # "本部"単位の各合計金額
            transportation_total_group += project_month.billing_transportation or 0
            tax_transportation_total_group += project_month.get_tax_of_billing_transportation()
            confirmation_total_group += project_month.billing_confirmation_money \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            tax_total_group += project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            amount_total_group += (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)
            # 全合計金額
            transportation_total_all += project_month.billing_transportation or 0
            tax_transportation_total_all += project_month.get_tax_of_billing_transportation()
            confirmation_total_all += project_month.billing_confirmation_money \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            tax_total_all += project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            amount_total_all += (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)

            group_name = project_month.project.recorded_department.group_name
            department_name = project_month.project.recorded_department.department_name
            company_name = project_month.project.client_company.company_name
            self.create_outline(i, 'department')
        # 最終合計行の追加
        self.ws['B' + str(i + self.current_row + 1)].value = company_name + " 合計"
        self.ws['B' + str(i + self.current_row + 2)].value = department_name + " 合計"
        self.ws['B' + str(i + self.current_row + 3)].value = group_name + " 合計"
        self.ws['B' + str(i + self.current_row + 4)].value = "総合計"
        # 値を代入
        self.ws['H' + str(i + self.current_row + 1)].value = transportation_total_company
        self.ws['I' + str(i + self.current_row + 1)].value = tax_transportation_total_company    
        self.ws['J' + str(i + self.current_row + 1)].value = confirmation_total_company
        self.ws['K' + str(i + self.current_row + 1)].value = tax_total_company
        self.ws['L' + str(i + self.current_row + 1)].value = amount_total_company
        self.ws['H' + str(i + self.current_row + 2)].value = transportation_total_department
        self.ws['I' + str(i + self.current_row + 2)].value = tax_transportation_total_department
        self.ws['J' + str(i + self.current_row + 2)].value = confirmation_total_department
        self.ws['K' + str(i + self.current_row + 2)].value = tax_total_department
        self.ws['L' + str(i + self.current_row + 2)].value = amount_total_department
        self.ws['H' + str(i + self.current_row + 3)].value = transportation_total_group
        self.ws['I' + str(i + self.current_row + 3)].value = tax_transportation_total_group
        self.ws['J' + str(i + self.current_row + 3)].value = confirmation_total_group
        self.ws['K' + str(i + self.current_row + 3)].value = tax_total_group
        self.ws['L' + str(i + self.current_row + 3)].value = amount_total_group
        self.ws['H' + str(i + self.current_row + 4)].value = transportation_total_all
        self.ws['I' + str(i + self.current_row + 4)].value = tax_transportation_total_all
        self.ws['J' + str(i + self.current_row + 4)].value = confirmation_total_all
        self.ws['K' + str(i + self.current_row + 4)].value = tax_total_all
        self.ws['L' + str(i + self.current_row + 4)].value = amount_total_all
        # セルの結合
        self.ws.merge_cells('B' + str(i + self.current_row + 1) + ':G' + str(i + self.current_row + 1))
        self.ws.merge_cells('B' + str(i + self.current_row + 2) + ':G' + str(i + self.current_row + 2))
        self.ws.merge_cells('B' + str(i + self.current_row + 3) + ':G' + str(i + self.current_row + 3))
        self.ws.merge_cells('B' + str(i + self.current_row + 4) + ':G' + str(i + self.current_row + 4))
        # セルの書式設定
        self.ws['B' + str(i + self.current_row + 1)].alignment = Alignment(horizontal='right')
        self.ws['B' + str(i + self.current_row + 2)].alignment = Alignment(horizontal='right')
        self.ws['B' + str(i + self.current_row + 3)].alignment = Alignment(horizontal='right')
        self.ws['B' + str(i + self.current_row + 4)].alignment = Alignment(horizontal='right')
        for column_num in ['B', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            # セルの色変更
            self.ws[column_num + str(i + self.current_row + 1)].fill = \
                PatternFill(patternType='solid', fgColor='B7D6A3')
            self.ws[column_num + str(i + self.current_row + 2)].fill = \
                PatternFill(patternType='solid', fgColor='94C175')
            self.ws[column_num + str(i + self.current_row + 3)].fill = \
                PatternFill(patternType='solid', fgColor='70AD47')
            self.ws[column_num + str(i + self.current_row + 4)].fill = \
                PatternFill(patternType='solid', fgColor='5A8A39')
            # フォントサイズの変更
            self.ws[column_num + str(i + self.current_row + 1)].font = Font(size=14, bold=True)
            self.ws[column_num + str(i + self.current_row + 2)].font = Font(size=14, bold=True)
            self.ws[column_num + str(i + self.current_row + 3)].font = Font(size=14, bold=True)
            self.ws[column_num + str(i + self.current_row + 4)].font = Font(size=14, bold=True)

        for j in range(5):
            self.money_format(i + j, 'department')
            self.create_outline(i + j, 'department')
        self.ws.title = '請求一覧（部）{}月'.format(self.month.month)

    def _create_excel_deposit(self):
        self.ws = self.excel.workbook['Sheet2']

        company_name = ""
        deposit_date = ""
        bank_name = ""
        # "部"単位の各合計金額の宣言
        transportation_total_department = 0
        tax_transportation_total_department = 0
        confirmation_total_department = 0
        tax_total_department = 0
        amount_total_department = 0
        # "全合計金額の宣言
        transportation_total_all = 0
        tax_transportation_total_all = 0
        confirmation_total_all = 0
        tax_total_all = 0
        amount_total_all = 0
        self.current_row = 3
        i = 0
        # 入金日、会社名順にソート
        self.project_months.sort(key=lambda x: (x.deposit_date, x.project.client_company.company_name))
        for i, project_month in enumerate(self.project_months):
            if i != 0:
                if company_name != project_month.project.client_company.company_name:
                    self.ws['B' + str(i + self.current_row)].value = company_name + " 集計"
                    # 値を代入
                    self.ws['C' + str(i + self.current_row)].value = transportation_total_department
                    self.ws['D' + str(i + self.current_row)].value = tax_transportation_total_department
                    self.ws['E' + str(i + self.current_row)].value = confirmation_total_department
                    self.ws['F' + str(i + self.current_row)].value = tax_total_department
                    self.ws['G' + str(i + self.current_row)].value = amount_total_department
                    self.ws['H' + str(i + self.current_row)].value = deposit_date
                    self.ws['I' + str(i + self.current_row)].value = bank_name
                    # 金額セルの書式設定
                    self.create_outline(i, 'deposit')
                    # セルの書式設定
                    self.money_format(i, 'deposit')
                    # 日付セルの書式設定
                    self.day_format(i, 'deposit')
                    # 合計金額の初期化
                    transportation_total_department = 0
                    tax_transportation_total_department = 0
                    confirmation_total_department = 0
                    tax_total_department = 0
                    amount_total_department = 0
                    self.current_row += 1
            # 値を代入
            self.ws['B' + str(i + self.current_row)].value = project_month.project.project_name
            self.ws['C' + str(i + self.current_row)].value = project_month.billing_transportation or 0
            self.ws['D' + str(i + self.current_row)].value = project_month.get_tax_of_billing_transportation()
            self.ws['E' + str(i + self.current_row)].value = (project_month.billing_confirmation_money or 0) \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            self.ws['F' + str(i + self.current_row)].value = project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            self.ws['G' + str(i + self.current_row)].value = (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)
            self.ws['H' + str(i + self.current_row)].value = project_month.deposit_date
            self.ws['I' + str(i + self.current_row)].value = project_month.project.client_company.bank.bank_name
            # セルの書式設定
            self.create_outline(i, 'deposit')
            self.ws.row_dimensions[i + self.current_row].hidden = True
            # 金額セルの書式設定
            self.money_format(i, 'deposit')
            # 日付セルの書式設定
            self.day_format(i, 'deposit')
            # "部"単位の各合計金額
            transportation_total_department += project_month.billing_transportation or 0
            tax_transportation_total_department += project_month.get_tax_of_billing_transportation()
            confirmation_total_department += project_month.billing_confirmation_money \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            tax_total_department += project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            amount_total_department += (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)
            # "全合計金額"の各合計金額
            transportation_total_all += project_month.billing_transportation or 0
            tax_transportation_total_all += project_month.get_tax_of_billing_transportation()
            confirmation_total_all += project_month.billing_confirmation_money \
                + (project_month.billing_transportation or 0) - project_month.get_tax_of_billing_transportation()
            tax_total_all += project_month.tax_of_billing_confirmation_money() \
                + project_month.get_tax_of_billing_transportation() 
            amount_total_all += (project_month.billing_confirmation_money or 0) \
                + project_month.tax_of_billing_confirmation_money() + (project_month.billing_transportation or 0)

            company_name = project_month.project.client_company.company_name
            deposit_date = project_month.deposit_date
            bank_name = project_month.project.client_company.bank.bank_name

        # 最終合計行の追加
        self.ws['B' + str(i + self.current_row + 1)].value = company_name + " 集計"
        self.ws['B' + str(i + self.current_row + 2)].value = "総計"
        # 値を代入
        self.ws['C' + str(i + self.current_row + 1)].value = transportation_total_department
        self.ws['D' + str(i + self.current_row + 1)].value = tax_transportation_total_department
        self.ws['E' + str(i + self.current_row + 1)].value = confirmation_total_department
        self.ws['F' + str(i + self.current_row + 1)].value = tax_total_department
        self.ws['G' + str(i + self.current_row + 1)].value = amount_total_department
        self.ws['H' + str(i + self.current_row + 1)].value = deposit_date
        self.ws['I' + str(i + self.current_row + 1)].value = bank_name
        self.ws['C' + str(i + self.current_row + 2)].value = transportation_total_all
        self.ws['D' + str(i + self.current_row + 2)].value = tax_transportation_total_all
        self.ws['E' + str(i + self.current_row + 2)].value = confirmation_total_all
        self.ws['F' + str(i + self.current_row + 2)].value = tax_total_all
        self.ws['G' + str(i + self.current_row + 2)].value = amount_total_all
        for j in range(3):
            # セルの書式設定
            self.create_outline(i + j, 'deposit')
            self.money_format(i + j, 'deposit')
            self.day_format(i + j, 'deposit')
        self.ws.title = '請求一覧（入金日）{}月'.format(self.month.month)
        # エクセルを一時フォルダに保存
        self.excel.save('請求一覧_{}.xlsx'.format(datetime.today().strftime("%Y%m%d")))

    def create_outline(self, index, kind):
        
        if kind == self.DEPARTMENT:
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                self.ws[col + str(index + self.current_row)].border = Border(outline=True,
                                                                             top=Side(style='thin'),
                                                                             left=Side(style='thin'),
                                                                             right=Side(style='thin'),
                                                                             bottom=Side(style='thin'))
        elif kind == self.DEPOSIT:
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                self.ws[col + str(index + self.current_row)].border = Border(outline=True,
                                                                             top=Side(style='thin'),
                                                                             left=Side(style='thin'),
                                                                             right=Side(style='thin'),
                                                                             bottom=Side(style='thin'))

    def day_format(self, index, kind):

        if kind == self.DEPARTMENT:
            for col in ['F', 'M']:
                self.ws[col + str(index + self.current_row)].number_format = 'yyyy/mm/dd'
        elif kind == self.DEPOSIT:
            self.ws['H' + str(index + self.current_row)].number_format = 'yyyy/mm/dd'

    def money_format(self, index, kind):

        if kind == self.DEPARTMENT:
            for col in ['H', 'I', 'J', 'K', 'L']:
                self.ws[col + str(index + self.current_row)].number_format = '¥#,###0'
        elif kind == self.DEPOSIT:
            for col in ['C', 'D', 'E', 'F', 'G']:
                self.ws[col + str(index + self.current_row)].number_format = '¥#,###0'

    def amount_total(self, index, col, name, transportation, tax_transportation, confirmation, tax_total, amount_total):

        self.ws['B' + str(index + self.current_row)].value = name + " 合計"
        # 値を代入
        self.ws['H' + str(index + self.current_row)].value = transportation
        self.ws['I' + str(index + self.current_row)].value = tax_transportation
        self.ws['J' + str(index + self.current_row)].value = confirmation
        self.ws['K' + str(index + self.current_row)].value = tax_total
        self.ws['L' + str(index + self.current_row)].value = amount_total
        # 金額セルの書式設定
        self.create_outline(index, 'department')
        # セルの結合
        self.ws.merge_cells('B' + str(index + self.current_row) + ':G' + str(index + self.current_row))
        # セルの書式設定
        self.ws['B' + str(index + self.current_row)].alignment = Alignment(horizontal='right')
        # セルの色変更
        for column_num in ['B', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            if col == 'company':
                self.ws[column_num + str(index + self.current_row)].fill = \
                    PatternFill(patternType='solid', fgColor='B7D6A3')
            elif col == 'department':
                self.ws[column_num + str(index + self.current_row)].fill = \
                    PatternFill(patternType='solid', fgColor='94C175')
            else:
                self.ws[column_num + str(index + self.current_row)].fill = \
                    PatternFill(patternType='solid', fgColor='70AD47')
        # フォントサイズの変更
        for column_num in ['B', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            self.ws[column_num + str(index + self.current_row)].font = Font(size=14, bold=True)

        self.create_outline(index, 'department')
        self.money_format(index, 'department')      
        self.current_row += 1
