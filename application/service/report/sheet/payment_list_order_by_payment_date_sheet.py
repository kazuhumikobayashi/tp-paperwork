from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

from application.domain.repository.engineer_history_repository import EngineerHistoryRepository

engineer_history_repository = EngineerHistoryRepository()


# 支払日ごとの支払一覧を作成する
class PaymentListByPaymentDateSheet(object):

    def __init__(self, payment_list_order_by_payment_date, month, ws):
        self.payment_list_order_by_payment_date = payment_list_order_by_payment_date
        self.month = month
        self.ws = ws
        self.current_row = 3

    def write_payment_list_order_by_payment_date(self):
        # 支払日ごとの支払実績、消費税、支払(税込)を計算
        total_money_excluding_tax_by_payment_date = 0
        total_tax_by_payment_date = 0
        total_transportation_by_payment_date = 0
        total_money_by_payment_date = 0
        # 支払実績、消費税、支払(税込)の総計を計算
        total_money_excluding_tax = 0
        total_tax = 0
        total_transportation = 0
        total_money = 0

        # 支払の月を記載
        self.ws.merge_cells('B1:L1')
        self.ws['B1'].value = self.month
        self.ws['B1'].number_format = 'yyyy"年"m"月度"'
        self.ws['B1'].font = Font(name="ＭＳ ゴシック", size=12, bold=True)
        self.ws['B1'].alignment = Alignment(horizontal="center")

        for i in range(len(self.payment_list_order_by_payment_date)):
            engineer_history = engineer_history_repository.get_history_at_result_month(
                                                self.payment_list_order_by_payment_date[i].project_detail.engineer_id,
                                                self.payment_list_order_by_payment_date[i].result_month)
            # 値を代入
            self.ws['B' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].project_detail.engineer.company.company_name
            self.ws['C' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].project_detail.engineer.engineer_name
            self.ws['D' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].project_detail.project.project_name
            self.ws['E' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].project_detail.engineer.company.payment_site.value
            if engineer_history:
                self.ws['F' + str(self.current_row)].value = engineer_history.payment_tax.name
            else:
                self.ws['F' + str(self.current_row)].value = ''
            self.ws['G' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].billing_receipted_date
            self.ws['H' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].payment_expected_date
            self.ws['I' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].payment_confirmation_money
            self.ws['J' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].tax_of_payment_confirmation_money(engineer_history)
            self.ws['K' + str(self.current_row)].value\
                = self.payment_list_order_by_payment_date[i].payment_transportation
            self.ws['L' + str(self.current_row)].value\
                = (self.payment_list_order_by_payment_date[i].payment_confirmation_money or 0)\
                + self.payment_list_order_by_payment_date[i].tax_of_payment_confirmation_money(engineer_history)\
                + (self.payment_list_order_by_payment_date[i].payment_transportation or 0)
            # 支払日ごとの支払実績、消費税、支払(税込)を計算
            total_money_excluding_tax_by_payment_date\
                += (self.payment_list_order_by_payment_date[i].payment_confirmation_money or 0)
            total_tax_by_payment_date\
                += self.payment_list_order_by_payment_date[i].tax_of_payment_confirmation_money(engineer_history)
            total_transportation_by_payment_date\
                += (self.payment_list_order_by_payment_date[i].payment_transportation or 0)
            total_money_by_payment_date\
                += (self.payment_list_order_by_payment_date[i].payment_confirmation_money or 0)\
                + self.payment_list_order_by_payment_date[i].tax_of_payment_confirmation_money(engineer_history)\
                + (self.payment_list_order_by_payment_date[i].payment_transportation or 0)

            # 罫線
            self.write_border_of_payment_list()
            # フォント
            for column_num in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                self.ws[column_num + str(self.current_row)].font = Font(name='ＭＳ ゴシック')
            # 書式設定
            for column_num in ['E', 'F', 'G', 'H']:
                self.ws[column_num + str(self.current_row)].alignment = Alignment(horizontal='center')
            self.ws['B' + str(self.current_row)].alignment = Alignment(shrink_to_fit=True)
            # ユーザー定義
            for column_num in ['G', 'H']:
                self.ws[column_num + str(self.current_row)].number_format = 'm/d'
            for column_num in ['I', 'J', 'K', 'L']:
                self.ws[column_num + str(self.current_row)].number_format = '#,##0'
            # 行の高さ調整
            self.ws.row_dimensions[self.current_row].height = 18
            # セルの背景
            for column_num in ['E', 'F']:
                self.ws[column_num + str(self.current_row)].fill = PatternFill(start_color='F6CA8D',
                                                                               end_color='F6CA8D',
                                                                               fill_type='solid')
            self.current_row += 1

            # 支払日ごとの最終行に小計を記載
            if (len(self.payment_list_order_by_payment_date) == (i + 1))\
                    or (self.payment_list_order_by_payment_date[i].payment_expected_date
                        != self.payment_list_order_by_payment_date[i+1].payment_expected_date):
                # セル結合
                self.ws.merge_cells('G' + str(self.current_row) + ':H' + str(self.current_row))
                self.ws['G' + str(self.current_row)].value\
                    = self.payment_list_order_by_payment_date[i].payment_expected_date
                self.ws['I' + str(self.current_row)].value = total_money_excluding_tax_by_payment_date
                self.ws['J' + str(self.current_row)].value = total_tax_by_payment_date
                self.ws['K' + str(self.current_row)].value = total_transportation_by_payment_date
                self.ws['L' + str(self.current_row)].value = total_money_by_payment_date
                # 支払実績、消費税、支払(税込)の総計を計算
                total_money_excluding_tax += total_money_excluding_tax_by_payment_date
                total_tax += total_tax_by_payment_date
                total_transportation += total_transportation_by_payment_date
                total_money += total_money_by_payment_date
                # 支払日ごとの消費税、支払(税込)をリセット
                total_money_excluding_tax_by_payment_date = 0
                total_tax_by_payment_date = 0
                total_transportation_by_payment_date = 0
                total_money_by_payment_date = 0
                # 書式設定
                self.ws['G' + str(self.current_row)].number_format = 'm/d  集計'
                self.ws['G' + str(self.current_row)].font = Font(bold=True)
                self.ws['G' + str(self.current_row)].alignment = Alignment(horizontal='right')
                for column_num in ['I', 'J', 'K', 'L']:
                    self.ws[column_num + str(self.current_row)].font = Font(name='ＭＳ ゴシック', bold=True)
                    self.ws[column_num + str(self.current_row)].alignment = Alignment(vertical='bottom')
                    self.ws[column_num + str(self.current_row)].number_format = '#,##0'
                # 罫線
                self.write_border_of_payment_list()
                # 行の高さ調整
                self.ws.row_dimensions[self.current_row].height = 18
                # セルの背景
                for column_num in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                    self.ws[column_num + str(self.current_row)].fill = PatternFill(start_color='00FFFF',
                                                                                   end_color='00FFFF',
                                                                                   fill_type='solid')
                self.current_row += 1

        # 総計
        # 値の代入
        self.ws['H' + str(self.current_row)].value = '総計'
        self.ws['I' + str(self.current_row)].value = total_money_excluding_tax
        self.ws['J' + str(self.current_row)].value = total_tax
        self.ws['K' + str(self.current_row)].value = total_transportation
        self.ws['L' + str(self.current_row)].value = total_money
        # 表示形式
        self.ws['H' + str(self.current_row)].alignment = Alignment(horizontal='center')
        for column_num in ['I', 'J', 'K', 'L']:
            self.ws[column_num + str(self.current_row)].number_format = '#,##0'
        # フォント
        self.ws['H' + str(self.current_row)].font = Font(bold=True)
        for column_num in ['I', 'J', 'K', 'L']:
            self.ws[column_num + str(self.current_row)].font = Font(name="ＭＳ ゴシック")
        # 罫線
        self.write_border_of_payment_list(top_border_style='double')
        # 行の高さ調整
        self.ws.row_dimensions[self.current_row].height = 18
        # セルの背景
        for column_num in ['E', 'F']:
            self.ws[column_num + str(self.current_row)].fill = PatternFill(start_color='F6CA8D',
                                                                           end_color='F6CA8D',
                                                                           fill_type='solid')

    def write_border_of_payment_list(self, top_border_style='thin'):
        for column_num in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            self.ws[column_num + str(self.current_row)].border = Border(top=Side(style=top_border_style),
                                                                        left=Side(style='medium'),
                                                                        right=Side(style='medium'),
                                                                        bottom=Side(style='thin'))
