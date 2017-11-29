from openpyxl.styles import Alignment, Border, Side, Font


# 部署ごとの支払一覧を作成する
from application.domain.repository.engineer_history_repository import EngineerHistoryRepository


class PaymentListByDepartmentSheet(object):
    engineer_history_repository = EngineerHistoryRepository()

    def __init__(self, payment_lists_by_department, month, ws):
        self.payment_lists_by_department = payment_lists_by_department
        self.month = month
        self.ws = ws
        self.current_row = 3

    def write_payment_list_by_department(self):
        # 支払の月を記載
        self.ws.merge_cells('B1:K1')
        self.ws['B1'].value = self.month
        self.ws['B1'].number_format = 'yyyy"年"m"月度"'
        self.ws['B1'].font = Font(name="ＭＳ ゴシック", size=12, bold=True)
        self.ws['B1'].alignment = Alignment(horizontal="center")

        # 部署ごとに分解
        for payment_list_by_department in self.payment_lists_by_department:
            # PRJ計、部署計
            prj_total_money = 0
            department_total_money = 0

            # 部署名を記入
            # BPの実績が2つ以上の場合、結合する
            if len(payment_list_by_department.project_results) >= 2:
                self.ws.merge_cells('B' + str(self.current_row) + ':B' + str(
                    self.current_row + len(payment_list_by_department.project_results) - 1))
            self.ws['B' + str(self.current_row)].value = payment_list_by_department.department.department_name[0:2]
            self.ws['B' + str(self.current_row)].alignment = Alignment(wrap_text=True,
                                                                       horizontal="center",
                                                                       vertical="center")

            # 顧客会社ごとに分解
            for i in range(len(payment_list_by_department.project_results)):
                engineer_history = self.engineer_history_repository.get_history_at_result_month(
                    payment_list_by_department.project_results[i].project_detail.engineer_id,
                    payment_list_by_department.project_results[i].result_month)
                # 値を代入
                self.ws['C' + str(self.current_row)].value = (i + 1)
                self.ws['D' + str(self.current_row)].value\
                    = payment_list_by_department.project_results[i].project_detail.engineer.company.company_name
                self.ws['E' + str(self.current_row)].value\
                    = payment_list_by_department.project_results[i].project_detail.project.client_company.company_name
                self.ws['F' + str(self.current_row)].value\
                    = payment_list_by_department.project_results[i].project_detail.engineer.engineer_name
                self.ws['G' + str(self.current_row)].value\
                    = payment_list_by_department.project_results[i].project_detail.project.project_name
                # 支払実績 = 税抜実績 ＋ 税抜交通費
                self.ws['H' + str(self.current_row)].value\
                    = (payment_list_by_department.project_results[i].payment_confirmation_money or 0)\
                    + (payment_list_by_department.project_results[i].payment_transportation or 0)\
                    - payment_list_by_department.project_results[i].get_tax_of_payment_transportation(engineer_history)
                # プロジェクト計の合計金額(税抜き)を取得
                prj_total_money += (payment_list_by_department.project_results[i].payment_confirmation_money or 0)\
                    + (payment_list_by_department.project_results[i].payment_transportation or 0)\
                    - payment_list_by_department.project_results[i].get_tax_of_payment_transportation(engineer_history)
                # 罫線（それぞれの部署の一行目には、上に太い罫線を引く）
                self.write_border()
                # 表示形式
                self.ws['C' + str(self.current_row)].alignment = Alignment(horizontal='center')
                # フォント
                for column_num in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    self.ws[column_num + str(self.current_row)].font = Font(name="ＭＳ ゴシック")
                # ユーザー定義
                for column_num in ['H', 'I', 'J']:
                    self.ws[column_num + str(self.current_row)].number_format = '#,##0'
                # 行の高さ調整
                self.ws.row_dimensions[self.current_row].height = 18
                # プロジェクトの最終行にはプロジェクト計を記載
                if len(payment_list_by_department.project_results) == (i + 1)\
                        or (payment_list_by_department.project_results[i].project_detail.project.project_name
                            != payment_list_by_department.project_results[i + 1].project_detail.project.project_name):
                    # PRJ計、部署計を代入
                    self.ws['I' + str(self.current_row)].value = prj_total_money
                    department_total_money += prj_total_money
                    # PRJ計リセット
                    prj_total_money = 0
                    # 罫線（それぞれの部署の一行目には、上に太い罫線を引く）
                    self.write_border(bottom_border_style='medium')
                self.current_row += 1

            # 計上しているプロジェクトが存在しない、もしくは計上しているプロジェクトにBPの実績がない場合
            if not payment_list_by_department.project_results:
                self.write_border_if_not_exist_bp()

            # 部署ごとの最終行の備考に部署ごとの合計を記載
            self.ws['J' + str(self.current_row - 1)].value = department_total_money

        # 合計------------------------------------------------------------------------------------------------------
        # セルを結合
        self.ws.merge_cells('B' + str(self.current_row) + ':G' + str(self.current_row))
        # 値を代入
        self.ws['B' + str(self.current_row)].value = "合計"
        self.ws['H' + str(self.current_row)].value = "=sum(H3:H{})".format(self.current_row - 1)
        self.ws['I' + str(self.current_row)].value = "=sum(I3:I{})".format(self.current_row - 1)
        self.ws['J' + str(self.current_row)].value = "=sum(J3:J{})".format(self.current_row - 1)
        # 書式設定
        # 書式
        self.ws['B' + str(self.current_row)].alignment = Alignment(horizontal="right")
        for column_num in ['H', 'I', 'J', 'K']:
            self.ws[column_num + str(self.current_row)].font = Font(name="ＭＳ ゴシック")
        # ユーザー定義
        for column_num in ['H', 'I', 'J']:
            self.ws[column_num + str(self.current_row)].number_format = '#,##0'
        # 行の高さ調整
        self.ws.row_dimensions[self.current_row].height = 18
        # 罫線
        self.write_border(bottom_border_style='medium')
        self.current_row += 1
        for column_num in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            self.ws[column_num + str(self.current_row)].border = Border(top=Side(style='medium'))

    def write_border(self, top_border_style=None, bottom_border_style='dotted'):
        for column_num in ['B']:
            self.ws[column_num + str(self.current_row)].border = Border(top=Side(style=top_border_style),
                                                                        left=Side(style='medium'),
                                                                        right=Side(style='medium'),
                                                                        bottom=Side(style=bottom_border_style))
        for column_num in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            self.ws[column_num + str(self.current_row)].border = Border(top=Side(style=top_border_style),
                                                                        left=Side(style='medium'),
                                                                        right=Side(style='medium'),
                                                                        bottom=Side(style=bottom_border_style))

    # BPがいないorその月の実績がない計上部署のスタイル処理
    def write_border_if_not_exist_bp(self):
        # 罫線
        self.write_border(bottom_border_style='medium')
        # 非表示にする。
        self.ws.row_dimensions[self.current_row].hidden = True
        self.current_row += 1
