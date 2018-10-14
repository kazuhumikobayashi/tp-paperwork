from datetime import datetime

from openpyxl.styles import Border, Side, Font, PatternFill

from application.domain.model.immutables.tax import Tax
from application.domain.repository.excel import Excel
from application.domain.repository.project_billing_repository import ProjectBillingRepository
from application.service.report.billing_base_report import BillingBaseReport


class BillingReport(BillingBaseReport):
    project_billing_repository = ProjectBillingRepository()

    def __init__(self, project_month):
        super(BillingReport, self).__init__(project_month)
        self.excel = Excel("billing.xlsx")
        self.ws = self.excel.active

    def billing_report_download(self):
        self._create_excel()

        return self.excel.download()

    def _create_excel(self):
        super(BillingReport, self)._create_excel()

        # 宛名シート作成
        self.address_sheet.create_address_sheet()

        # エクセルを一時フォルダに保存
        self.excel.save('06_請求書（{}）_{}.xlsx'.format(self.project_month.client_billing_no,
                                                    datetime.today().strftime("%Y%m%d")))

    def write_top_part(self):
        super(BillingReport, self).write_top_part()

        # フォントを調整
        self.ws[self.excel.get_defined_name_range("total_money_title")].font = \
            Font(name="HGP明朝", size=14, underline="single")
        self.ws[self.excel.get_defined_name_range("total_money")].font = \
            Font(name="Century", size=14, underline="single")

    def create_billing_details(self):
        project_billings = self.project_billing_repository.find_billings_at_a_month(self.project_month.project_id,
                                                                                    self.project_month.project_month)
        # 請求明細----------------------------------------------------------------------------------
        for n, project_billing in enumerate(project_billings):
            # 書式設定
            self.create_billing_detail()
            # 値代入
            self.ws['A' + str(self.current_row)].value = (n + 1)
            self.ws['E' + str(self.current_row)].value = project_billing.billing_content
            self.ws['I' + str(self.current_row)].value = project_billing.billing_amount
            self.ws['K' + str(self.current_row)].value = project_billing.billing_confirmation_money
            self.ws['M' + str(self.current_row)].value = project_billing.remarks
            # 次の行へ移動
            self.current_row += 1
        self.current_row += 1

        # 行調整
        if self.current_row < 25:
            self.current_row = 25

        # 小計--------------------------------------------------------------------------------------
        # 値代入
        self.create_subtotal_list('課税　小計', self.project_month.billing_confirmation_money)
        if self.project_month.project.client_company.billing_tax != Tax.zero:
            self.create_subtotal_list('消費税', self.project_month.tax_of_billing_confirmation_money())
        self.create_subtotal_list('交通費等　非課税　小計', self.project_month.billing_transportation)
        self.current_row += 1

        # 行調整
        if self.current_row < 39:
            self.current_row = 39

        # 累計--------------------------------------------------------------------------------------
        # 値代入
        self.ws['I' + str(self.current_row)].value = '計'
        self.ws['K' + str(self.current_row)].value = (self.project_month.billing_confirmation_money or 0)\
            + self.project_month.tax_of_billing_confirmation_money()\
            + (self.project_month.billing_transportation or 0)
        self.ws['K14'].value = '=K' + str(self.current_row)
        # 書式設定
        self.ws['I' + str(self.current_row)].font = Font(name="HGS明朝")
        self.ws['K' + str(self.current_row)].font = Font(name="Century")
        # 罫線
        for row in self.ws['A' + str(self.current_row) + ":Q" + str(self.current_row)]:
            for cell in row:
                cell.border = Border(top=Side(style='thin'))
        # セルの色変更
        for column_num in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
            self.ws[column_num + str(self.current_row)].fill = PatternFill(patternType='solid', fgColor='E8F0F8')
        self.current_row += 2

    def create_bottom_part(self):
        super(BillingReport, self).create_bottom_part()

        # お支払先記入
        # 値代入
        self.ws['B' + str(self.current_row)].value = self.project_month.project.client_company.bank.text_for_document
