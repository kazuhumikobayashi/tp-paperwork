import locale

from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

from application.service.report.sheet.address_sheet import AddressSheet


class BillingBaseReport(object):

    def __init__(self, project_month):
        self.project_month = project_month
        self.excel = None
        self.ws = None
        self.current_row = 19
        self.address_sheet = AddressSheet(self)

    def _create_excel(self):
        # 17行目までの値代入・書式設定
        self.write_top_part()

        # 請求明細部分を作成
        self.create_billing_details()

        # 備考・お支払先部分を作成
        self.create_bottom_part()

        # ロゴ画像貼り付け
        img = Image('excel/templates/tp_logo.png', size=(290, 100))
        self.ws.add_image(img, 'K7')

        # 倍率の調整
        self.ws.page_setup.scale = 85

    def write_top_part(self):
        # 値を代入
        self.ws.get_named_range("client_billing_no")[0].value = self.project_month.client_billing_no
        self.ws.get_named_range("printed_date")[0].value = self.project_month.billing_printed_date
        self.ws.get_named_range("client_company_name")[0].value = \
            self.project_month.project.client_company.company_name
        self.ws.get_named_range("project_name")[0].value = self.project_month.project.project_name

        # フォント
        self.ws.get_named_range("title")[0].font = Font(name="HGP明朝", size=19, bold=True, underline="single")
        self.ws.get_named_range("printed_date")[0].font = Font(name="HGP明朝", underline="single")
        self.ws.get_named_range("client_company_name")[0].font = Font(name="HGP明朝", size=13, underline="single")
        self.ws.get_named_range("project_name_title")[0].font = Font(name="HGP明朝", size=13, underline="single")
        self.ws.get_named_range("project_name")[0].font = Font(name="HGP明朝", size=13, underline="single")

        # 表示形式
        self.ws.get_named_range("printed_date")[0].number_format = 'yyyy年 m月 d日'

        # 罫線
        for column_num in ['N', 'O']:
            self.ws[column_num + str(13)].border = Border(top=Side(style='thin'),
                                                          left=Side(style='thin'),
                                                          right=Side(style='thin'))
            self.ws[column_num + str(14)].border = Border(left=Side(style='thin'), right=Side(style='thin'))
            self.ws[column_num + str(15)].border = Border(left=Side(style='thin'),
                                                          right=Side(style='thin'),
                                                          bottom=Side(style='thin'))

    def create_billing_details(self):
        pass

    def create_billing_detail(self):
        # セルの結合
        self.ws.merge_cells('A' + str(self.current_row) + ':D' + str(self.current_row))
        self.ws.merge_cells('E' + str(self.current_row) + ':H' + str(self.current_row))
        self.ws.merge_cells('I' + str(self.current_row) + ':J' + str(self.current_row))
        self.ws.merge_cells('M' + str(self.current_row) + ':Q' + str(self.current_row))
        # フォント設定
        self.ws['A' + str(self.current_row)].font = Font(name="Century")
        self.ws['E' + str(self.current_row)].font = Font(name="HGS明朝")
        self.ws['I' + str(self.current_row)].font = Font(name="HGS明朝")
        self.ws['K' + str(self.current_row)].font = Font(name="Century")
        self.ws['M' + str(self.current_row)].font = Font(name="HGS明朝")
        # 書式設定
        self.ws['A' + str(self.current_row)].alignment = Alignment(horizontal='center', vertical='top')
        self.ws['E' + str(self.current_row)].alignment = Alignment(wrap_text=True, vertical='top')
        self.ws['I' + str(self.current_row)].alignment = Alignment(vertical='top')
        self.ws['K' + str(self.current_row)].alignment = Alignment(vertical='top')
        self.ws['M' + str(self.current_row)].alignment = Alignment(wrap_text=True, vertical='top')
        # 行調整（2行）
        self.ws.row_dimensions[self.current_row].height = 13.2 * 2
        # ユーザー定義
        self.ws['K' + str(self.current_row)].number_format = '#,##0'
        # 偶数行の色を変更
        if self.current_row % 2 == 1:
            for column_num in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
                self.ws[column_num + str(self.current_row)].fill = PatternFill(patternType='solid',
                                                                               fgColor='E8F0F8')

    def create_subtotal_list(self, title, value):
        # 値代入
        self.ws['E' + str(self.current_row)].value = title
        if title == '消費税':
            self.ws['I' + str(self.current_row)].value = self.project_month.billing_tax.name
        self.ws['K' + str(self.current_row)].value = value
        # 書式設定
        self.ws['E' + str(self.current_row)].font = Font(name="HGS明朝")
        self.ws['I' + str(self.current_row)].font = Font(name="HGS明朝")
        self.ws['K' + str(self.current_row)].font = Font(name="Century")
        # ユーザー定義
        self.ws['K' + str(self.current_row)].number_format = '#,##0'
        # 罫線
        for row in self.ws.iter_rows('A' + str(self.current_row) + ":L" + str(self.current_row)):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        self.current_row += 1

    def create_bottom_part(self):
        # 備考記入-------------------------------------------------------------------------------------
        # 行調整
        self.ws.row_dimensions[self.current_row].height = 9
        self.create_remark_top_border()
        self.ws.row_dimensions[self.current_row].height = 9
        self.create_remark_side_border()

        # セルの結合
        self.ws.merge_cells('D' + str(self.current_row) + ':P' + str(self.current_row + 15))
        # 値代入
        self.ws['B' + str(self.current_row)].value = '備考'
        self.ws['D' + str(self.current_row)].value = self.remark()
        # フォント
        self.ws['B' + str(self.current_row)].font = Font(name="HGS明朝", underline="single")
        self.ws['D' + str(self.current_row)].font = Font(name="HGS明朝")
        # 書式設定
        self.ws['B' + str(self.current_row)].alignment = Alignment(vertical='top')
        self.ws['D' + str(self.current_row)].alignment = Alignment(wrap_text=True, vertical='top')
        # 罫線
        for i in range(16):
            self.create_remark_side_border()
        self.create_remark_bottom_border()

    def create_remark_top_border(self):
        # 上部の罫線
        for row in self.ws.iter_rows('B' + str(self.current_row) + ":P" + str(self.current_row)):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'))
        self.current_row += 1

    def create_remark_side_border(self):
        # サイドの罫線
        self.ws['A' + str(self.current_row)].border = Border(right=Side(style='thin'))
        self.ws['Q' + str(self.current_row)].border = Border(left=Side(style='thin'))
        self.current_row += 1

    def create_remark_bottom_border(self):
        # 下部の罫線
        for row in self.ws.iter_rows('B' + str(self.current_row) + ":P" + str(self.current_row)):
            for cell in row:
                cell.border = Border(top=Side(style='thin'))
        self.current_row += 1

    def remark(self):
        if self.project_month.deposit_date is None:
            deposit_date = "               "
        else:
            locale.setlocale(locale.LC_ALL, '')
            deposit_date = self.project_month.deposit_date.strftime('%Y年%m月%d日')

        text = "お支払いは下記口座に{}までにお支払い願います。\n\n".format(deposit_date)\
               + "  ※作業内容及び納品物等の詳細はご注文書（No.{}）の通り"\
                 .format(self.project_month.project.client_order_no or "               ")
        return text

    def get_print_name(self):
        return self.project_month.project.client_company.print_name
