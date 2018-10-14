from datetime import datetime

from openpyxl.styles import Border, Side, Font

from application.domain.model.immutables.tax import Tax
from application.domain.repository.excel import Excel
from application.domain.repository.project_billing_repository import ProjectBillingRepository
from application.service.report.billing_base_report import BillingBaseReport


class DeliveryReport(BillingBaseReport):
    project_billing_repository = ProjectBillingRepository()

    def __init__(self, project_month):
        super(DeliveryReport, self).__init__(project_month)
        self.excel = Excel("delivery.xlsx")
        self.ws = self.excel.active

    def delivery_report_download(self):
        self._create_excel()

        return self.excel.download()

    def _create_excel(self):
        super(DeliveryReport, self)._create_excel()

        # エクセルを一時フォルダに保存
        self.excel.save('05_納品書（{}）_{}.xlsx'.format(self.project_month.client_billing_no,
                                                    datetime.today().strftime("%Y%m%d")))

    def create_billing_details(self):
        project_billings = self.project_billing_repository.find_billings_at_a_month(self.project_month.project_id,
                                                                                    self.project_month.project_month)
        # 請求明細----------------------------------------------------------------------------------
        for n, project_billing in enumerate(project_billings):
            # 書式設定
            self.create_billing_detail()
            # 値代入
            self.ws['A' + str(self.current_row)].value = (n + 1)
            self.ws['E' + str(self.current_row)].value = project_billing.billing_content
            self.ws['I' + str(self.current_row)].value = project_billing.billing_amount
            self.ws['M' + str(self.current_row)].value = project_billing.remarks
            # 次の行へ移動
            self.current_row += 1
        self.current_row += 1

        # 行調整
        if self.current_row < 25:
            self.current_row = 25

        # 小計--------------------------------------------------------------------------------------
        # 値代入
        self.create_subtotal_list('課税　小計', "")
        if self.project_month.project.client_company.billing_tax != Tax.zero:
            self.create_subtotal_list('消費税', "")
        self.create_subtotal_list('交通費等　非課税　小計', "")
        self.current_row += 1

        # 行調整
        if self.current_row < 39:
            self.current_row = 39

        # 累計--------------------------------------------------------------------------------------
        # 書式設定
        self.ws['I' + str(self.current_row)].font = Font(name="HGS明朝")
        self.ws['K' + str(self.current_row)].font = Font(name="Century")
        # 罫線
        for row in self.ws['A' + str(self.current_row) + ":Q" + str(self.current_row)]:
            for cell in row:
                cell.border = Border(top=Side(style='thin'))
        self.current_row += 2
