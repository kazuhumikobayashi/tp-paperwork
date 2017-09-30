from datetime import datetime

from openpyxl.styles import Alignment, Border, Side

from application.domain.model.immutables.rule import Rule
from application.domain.model.point import Point
from application.domain.repository.excel import Excel


class ProjectList(object):

    def __init__(self, project_details_bp, project_details_our_company, month):
        self.project_details_bp = project_details_bp
        self.project_details_our_company = project_details_our_company
        self.month = month
        self.excel = Excel("project_list.xlsx")
        self.current_row = 4
        self.current_col = 1
        self.ws = self.excel.active

    def download(self):
        self._create_excel_bp()
        self._create_excel_our_company()

        return self.excel.download()

    def _create_excel_bp(self):
        self.ws = self.excel.workbook['案件一覧(BP)']
        self._create_excel()

    def _create_excel_our_company(self):
        self.ws = self.excel.workbook['案件一覧(プロパー)']
        self._create_excel()

    def _create_excel(self):
        to = ""
        client_flag = ""
        if self.ws.title == '案件一覧(BP)':
            to = self.project_details_bp
            client_flag = 'BP'
        elif self.ws.title == '案件一覧(プロパー)':
            to = self.project_details_our_company
            client_flag = 'プロパー'

        # 結合しているセルの罫線が消えるので再度引き直す
        if client_flag == 'BP':
            for column_num in ['P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']:
                self.ws[column_num + '2'].border = Border(top=Side(style='thin'),
                                                          left=Side(style='thin'),
                                                          right=Side(style='thin'),
                                                          bottom=Side(style='thin'))
        else:
            for column_num in ['P', 'Q', 'R', 'S', 'T', 'U']:
                self.ws[column_num + '2'].border = Border(top=Side(style='thin'),
                                                          left=Side(style='thin'),
                                                          right=Side(style='thin'),
                                                          bottom=Side(style='thin'))           

        # セルにデータを書き込み
        i = 0
        for i, project_detail in enumerate(to):
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = ""
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = ""
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = ""
            if project_detail.project.recorded_department:
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail.project\
                    .recorded_department.department_name
            else:
                self.increment_col()
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail.project\
                .sales_person
            if project_detail.project.client_company:
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                    .project.client_company.company_name
            else:
                self.increment_col()
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .project.project_name
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .engineer.engineer_name
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .engineer.engineer_name_kana
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .billing_per_month
            if project_detail.project.client_company and project_detail.project.client_company.billing_site:
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail.project\
                    .client_company.billing_site.name
            else:
                self.increment_col()
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .engineer.company.company_name
            if project_detail.engineer.engineer_histories:
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail.engineer\
                    .engineer_histories[0].payment_per_month
            else:
                self.increment_col()
            if client_flag == 'BP':
                if project_detail.engineer.company and project_detail.engineer.company.payment_site:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .engineer.company.payment_site.name
                else:
                    self.increment_col()
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = "=J"\
                + str(i + self.current_row) + "-M" + str(i + self.current_row)
            if project_detail.billing_rule == Rule.fixed:
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                    .billing_rule.name
            else:
                if project_detail.billing_bottom_base_hour:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] =\
                        str(project_detail.billing_bottom_base_hour) + "～" + str(project_detail.billing_top_base_hour)
                else:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .billing_free_base_hour     
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .billing_per_hour
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .billing_per_bottom_hour
            if client_flag == 'BP':
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                    .billing_per_top_hour
                if project_detail.engineer.engineer_histories:
                    if project_detail.engineer.engineer_histories[0].payment_rule == Rule.fixed:
                        self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] =\
                            project_detail.engineer.engineer_histories[0].payment_rule.name
                    else:
                        if project_detail.engineer.engineer_histories[0].payment_bottom_base_hour:
                            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] =\
                                str(project_detail.engineer.engineer_histories[0].payment_bottom_base_hour)\
                                + "～" + str(project_detail.engineer.engineer_histories[0].payment_top_base_hour)
                        else:
                            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = \
                                project_detail.engineer.engineer_histories[0].payment_free_base_hour
                else:
                    self.increment_col()
                if project_detail.engineer.engineer_histories:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .engineer.engineer_histories[0].payment_per_hour
                else:
                    self.increment_col()
                if project_detail.engineer.engineer_histories:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .engineer.engineer_histories[0].payment_per_bottom_hour
                else:
                    self.increment_col()
                if project_detail.engineer.engineer_histories:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .engineer.engineer_histories[0].payment_per_top_hour
                else:
                    self.increment_col()
                if project_detail.engineer.engineer_histories:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .billing_start_day
                else:
                    self.increment_col()
                if project_detail.engineer.engineer_histories:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .billing_end_day
                else:
                    self.increment_col()
                if project_detail.engineer.engineer_histories:
                    self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                        .bp_order_no
                else:
                    self.increment_col()
            else:
                self.increment_col()
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .project.start_date
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .project.end_date
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .project.estimation_no
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = ""
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = ""
            if project_detail.project.end_user_company:
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                    .project.end_user_company.company_name
            else:
                self.increment_col()
            if project_detail.engineer.engineer_business_categories:
                self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail.engineer\
                    .engineer_business_categories[0].business_category.business_category_name
            else:
                self.increment_col()
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .engineer.get_age()
            skill = ''
            if project_detail.engineer.engineer_skills:
                skills = [engineer_skill.skill.skill_name for engineer_skill in project_detail.engineer.engineer_skills]
                skill = ','.join(skills)
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = skill
            self.ws[Point(i + self.current_row, self.increment_col()).get_cell_name()] = project_detail\
                .project.project_name_for_bp
 
            if client_flag == 'BP':
                # 罫線の設定
                for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
                            'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC',
                            'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']:
                    self.ws[col + str(i + self.current_row)].border = Border(outline=True,
                                                                             top=Side(style='dotted'),
                                                                             left=Side(style='thin'),
                                                                             right=Side(style='thin'),
                                                                             bottom=Side(style='dotted'))
                # 日付の設定
                for col in ['X', 'Y', 'AA', 'AB']:
                    self.ws[col + str(i + self.current_row)].number_format = 'yyyy/mm/dd'
     
                # 金額の設定
                for col in ['J', 'M', 'O']:
                    self.ws[col + str(i + self.current_row)].number_format = '¥#,###0'
     
                # セルの設定
                for col in ['K', 'N', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'Z', 'AC']:
                    self.ws[col + str(i + self.current_row)].alignment = Alignment(horizontal='center')
            else:
                # 罫線の設定
                for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
                            'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']:
                    self.ws[col + str(i + self.current_row)].border = Border(outline=True,
                                                                             top=Side(style='dotted'),
                                                                             left=Side(style='thin'),
                                                                             right=Side(style='thin'),
                                                                             bottom=Side(style='dotted'))
                # 日付の設定
                for col in ['S', 'T']:
                    self.ws[col + str(i + self.current_row)].number_format = 'yyyy/mm/dd'
     
                # 金額の設定
                for col in ['J', 'M', 'N']:
                    self.ws[col + str(i + self.current_row)].number_format = '¥#,###0'
     
                # セルの設定
                for col in ['K', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                    self.ws[col + str(i + self.current_row)].alignment = Alignment(horizontal='center')
            self.current_col = 1

        # 最終行の罫線の設定
        if client_flag == 'BP':
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
                        'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC',
                        'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']:
                self.ws[col + str(i + self.current_row)].border = Border(outline=True,
                                                                         top=Side(style='dotted'),
                                                                         left=Side(style='thin'),
                                                                         right=Side(style='thin'),
                                                                         bottom=Side(style='thin'))
        else:
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
                        'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']:
                self.ws[col + str(i + self.current_row)].border = Border(outline=True,
                                                                         top=Side(style='dotted'),
                                                                         left=Side(style='thin'),
                                                                         right=Side(style='thin'),
                                                                         bottom=Side(style='thin'))
        
        self.ws.title = self.ws.title + '{}月'.format(self.month.month)          

        # エクセルを一時フォルダに保存
        self.excel.save('案件一覧_{}.xlsx'.format(datetime.today().strftime("%Y%m%d")))
        
    def increment_col(self):
        col = self.current_col
        self.current_col += 1
        return col
